"""
fetch_data.py — Makro BiH
"""
import requests, json, os, re
try:
    import pdfplumber
except ImportError:
    os.system('pip install pdfplumber')
    import pdfplumber
from datetime import datetime
from io import BytesIO
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
    'Referer': 'https://bhas.gov.ba/',
}
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
os.makedirs(DATA_DIR, exist_ok=True)

def save_json(filename, data):
    path = os.path.join(DATA_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    print(f"  OK {filename} ({os.path.getsize(path)//1024} KB)")

def fetch_excel(url, timeout=30):
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    ct = r.headers.get('Content-Type', '')
    if 'html' in ct.lower(): raise ValueError(f"HTML umjesto Excel")
    if len(r.content) < 2000: raise ValueError(f"Premali ({len(r.content)} bytes)")
    return BytesIO(r.content)

def parse_num(val):
    if val is None: return None
    s = str(val).strip().replace('\xa0','').replace(' ','')
    if s in ('','-',':','...','n/a','N/A','x','X'): return None
    if re.match(r'^-?[\d.]+,\d+$', s): s = s.replace('.','').replace(',','.')
    try: return round(float(s), 2)
    except: return None

def is_period(val):
    s = str(val or '').strip()
    return bool(re.match(r'^(19|20)\d{2}$', s) or
                re.match(r'^(19|20)\d{2}-\d{1,2}$', s) or
                re.match(r'^(19|20)\d{2}[Qq]\d$', s))

def sort_periods(periods):
    def norm(p):
        if '-' in p:
            yr, mo = p.split('-', 1)
            return yr + '-' + mo.zfill(2)
        return p
    return sorted(periods, key=norm)

def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {}
    first_row = [str(c or '').strip() for c in rows[0]]
    periods_in_cols = sum(1 for c in first_row if is_period(c))
    result = {}
    if periods_in_cols >= 3:
        for row in rows[1:]:
            if not row[0]: continue
            label = str(row[0]).strip()
            if not label: continue
            series = {}
            for j, h in enumerate(first_row[1:], 1):
                if h and is_period(h) and j < len(row):
                    v = parse_num(row[j])
                    if v is not None: series[h] = v
            if series: result[label] = series
    else:
        col_headers = first_row[1:]
        for row in rows[1:]:
            if not row[0] or not is_period(str(row[0])): continue
            period = str(row[0]).strip()
            for j, h in enumerate(col_headers):
                if not h: continue
                v = parse_num(row[j+1] if j+1 < len(row) else None)
                if v is not None:
                    if h not in result: result[h] = {}
                    result[h][period] = v
    return result

def compact_series(series_dict, n_periods=72, max_series=20):
    if not series_dict: return {}
    all_periods = set()
    for s in series_dict.values(): all_periods.update(s.keys())
    periods = sort_periods(all_periods)[-n_periods:]
    totals = {k: sum(abs(v) for v in s.values()) for k, s in series_dict.items()}
    top_keys = sorted(totals, key=totals.get, reverse=True)[:max_series]
    return {k: {p: series_dict[k][p] for p in periods if p in series_dict[k]} for k in top_keys}

def fetch_standard(key, name, urls, sheets_idx, outfile, n_periods=72):
    print(f"-> {name}...")
    last_error = None
    for url in urls:
        fname = url.split('/')[-1]
        try:
            print(f"  Probam: {fname}")
            xls = fetch_excel(url)
            wb = openpyxl.load_workbook(xls, data_only=True)
            all_data = {}
            for idx in sheets_idx:
                if idx >= len(wb.sheetnames): continue
                sname = wb.sheetnames[idx]
                parsed = parse_sheet(wb[sname])
                if parsed:
                    compact = compact_series(parsed, n_periods=n_periods)
                    all_data[sname] = compact
                    first = next(iter(compact.values()))
                    print(f"    OK '{sname}': {len(compact)} serija, {len(first)} perioda")
            if not all_data: raise ValueError("Nema podataka")
            save_json(outfile, {'source':'BHAS','name':name,'url':url,
                                'updated':datetime.now().isoformat()[:10],'sheets':all_data})
            return True
        except Exception as e:
            print(f"  X {fname}: {e}")
            last_error = str(e)
    path = os.path.join(DATA_DIR, outfile)
    if not os.path.exists(path):
        save_json(outfile, {'source':'BHAS','name':name,'error':last_error,
                            'updated':datetime.now().isoformat()[:10],'sheets':{}})
    return False

def fetch_vanjska_trgovina():
    """
    ETR_01 = ukupna razmjena (izvoz+uvoz) po HS poglavljima
    ETR_02 = izvoz po HS poglavljima
    ETR_03 = uvoz po HS poglavljima
    """
    print("-> Vanjska trgovina (BHAS ETR_01/02/03)...")
    
    def load_etr(urls):
        for url in urls:
            fname = url.split('/')[-1]
            try:
                print(f"  Probam: {fname}")
                xls = fetch_excel(url)
                wb = openpyxl.load_workbook(xls, data_only=True)
                sname = wb.sheetnames[0]
                full = parse_sheet(wb[sname])
                if not full: raise ValueError("Nema podataka")
                print(f"    OK '{sname}': {len(full)} serija")
                return full, url
            except Exception as e:
                print(f"  X {fname}: {e}")
        return None, None

    # Učitaj sve tri serije
    etr01, url01 = load_etr(['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/ETR_01.xlsx'])
    etr02, url02 = load_etr(['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/ETR_02.xlsx'])
    etr03, url03 = load_etr(['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/ETR_03.xlsx'])

    if not etr01:
        save_json('vanjska_trgovina.json', {'source':'BHAS','error':'ETR_01 nije dostupan',
                                             'updated':datetime.now().isoformat()[:10],'data':{}})
        return False

    # Periodi - zadnjih 84
    all_periods = set()
    for s in etr01.values(): all_periods.update(s.keys())
    periods = sort_periods(all_periods)[-84:]

    # ETR_01 UK = IZVOZ BiH po HS poglavljima (potvrđeno s BHAS PDF saopštenjima)
    # ETR_02 UK = UVOZ BiH po zemljama porijekla (potvrđeno s BHAS PDF saopštenjima)
    # ETR_03 = nedostupan (403)
    
    izvoz_series = etr01.get('UK', {})
    izvoz = {p: izvoz_series[p] for p in periods if p in izvoz_series}
    print(f"  Izvoz (ETR_01 UK): {len(izvoz)} perioda")

    uvoz = {}
    if etr02:
        # ETR_02 ima UK seriju koja je ukupni uvoz
        uvoz_series = etr02.get('UK', etr02.get(list(etr02.keys())[0], {}))
        uvoz = {p: uvoz_series[p] for p in periods if p in uvoz_series}
        print(f"  Uvoz (ETR_02 UK): {len(uvoz)} perioda")
    
    # Za kompatibilnost sa starim kodom
    uk_compact = izvoz
    ex = izvoz
    im = uvoz

    # Godišnji zbroj
    annual_uk, annual_ex, annual_im = defaultdict(float), defaultdict(float), defaultdict(float)
    for p, v in izvoz.items():
        annual_uk[p.split('-')[0]] += v
    for p, v in ex.items():
        annual_ex[p.split('-')[0]] += v
    for p, v in im.items():
        annual_im[p.split('-')[0]] += v

    # Top 10 HS poglavlja po ukupnoj vrijednosti
    hs_totals = {k: sum(abs(v) for v in s.values()) 
                for k, s in etr01.items() if k != 'UK' and k.isdigit()}
    top10_hs = sorted(hs_totals, key=hs_totals.get, reverse=True)[:10]

    data = {
        'UK': uk_compact,
        'EX': izvoz,  # izvoz (ETR_01)
        'IM': uvoz,   # uvoz (ETR_02)
        'annual': {
            yr: {
                'uk': round(annual_uk.get(yr, 0)/1e9, 3),
                'ex': round(annual_ex.get(yr, 0)/1e9, 3),
                'im': round(annual_im.get(yr, 0)/1e9, 3),
            }
            for yr in sorted(set(list(annual_uk.keys())[-10:]))
        },
    }
    for k in top10_hs:
        data[k] = {p: etr01[k][p] for p in periods if p in etr01.get(k, {})}

    save_json('vanjska_trgovina.json', {
        'source': 'BHAS ETR_01/02/03',
        'name': 'Vanjska trgovina - izvoz, uvoz, razmjena BiH',
        'url': url01,
        'updated': datetime.now().isoformat()[:10],
        'note': 'EX=izvoz(ETR_01 UK), IM=uvoz(ETR_02 UK), 01-99=HS poglavlja',
        'has_ex': bool(izvoz),
        'has_im': bool(uvoz),
        'periods': periods,
        'data': data
    })
    return True

def fetch_cpi_pdf():
    """Parsira BHAS PDF saopstenja za CPI - PRI_01_YYYY_MM_1_BS.pdf"""
    print("-> CPI inflacija (BHAS PDF PRI_01)...")
    from datetime import date
    import re as re2
    today = date.today()
    cpi_data = {}

    for delta in range(0, 36):
        month = today.month - delta
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        period = str(year) + "-" + str(month)

        for lang in ["BS", "HR"]:
            url = f"https://bhas.gov.ba/data/Publikacije/Saopstenja/{year}/PRI_01_{year}_{month:02d}_1_{lang}.pdf"
            try:
                r = requests.get(url, headers=HEADERS, timeout=15)
                if r.status_code != 200 or len(r.content) < 3000:
                    continue
                with pdfplumber.open(BytesIO(r.content)) as pdf:
                    text = "".join(p.extract_text() or "" for p in pdf.pages[:2])
                if not text: continue
                for line in text.split("\n"):
                    if any(kw in line.upper() for kw in ["UKUPNO", "TOTAL", "CPI", "INDEKS"]):
                        nums = re2.findall(r"-?\d+\.\d+", line)
                        if len(nums) >= 2:
                            try:
                                vals = [float(n) for n in nums]
                                idx_c = [v for v in vals if 90 <= v <= 200]
                                if idx_c:
                                    cpi_data[period] = {"index": idx_c[0], "yoy": None}
                                    print(f"  OK {period} ({lang}): idx={idx_c[0]}")
                                    break
                            except: pass
                if period in cpi_data: break
            except Exception: pass

    if not cpi_data:
        print("  Nema CPI podataka")
        path = os.path.join(DATA_DIR, "cpi.json")
        if not os.path.exists(path):
            save_json("cpi.json", {"source":"BHAS","error":"PDF nije dostupan","updated":datetime.now().isoformat()[:10],"data":{}})
        return False

    sorted_p = sort_periods(list(cpi_data.keys()))
    save_json("cpi.json", {"source":"BHAS PRI_01","name":"CPI BiH","updated":datetime.now().isoformat()[:10],"periods":sorted_p,"data":{p:cpi_data[p] for p in sorted_p}})
    return True


def fetch_place_neto_bruto():
    """
    Parsira BHAS LAB_04 PDF saopstenja za prosjecne place u BiH.
    Format: "iznosi 1 684 KM" - broj je split u dva dijela razmakom.
    PDF takodje sadrzi podatke po svim sektorima za tekuci i prethodne godine.
    """
    print("-> Place neto i bruto (BHAS LAB_04 PDF + LAB_01 Excel)...")
    from datetime import date
    import re as re2
    today = date.today()

    place_data = {}

    for delta in range(0, 36):
        month = today.month - delta
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        period = str(year) + "-" + str(month)

        for lang in ["BS", "HR"]:
            url = f"https://bhas.gov.ba/data/Publikacije/Saopstenja/{year}/LAB_04_{year}_{month:02d}_1_{lang}.pdf"
            try:
                r = requests.get(url, headers=HEADERS, timeout=15)
                if r.status_code != 200 or len(r.content) < 3000: continue

                with pdfplumber.open(BytesIO(r.content)) as pdf:
                    text = "".join(p.extract_text() or "" for p in pdf.pages[:2])
                if not text: continue

                entry = {}

                # Trazi pattern "iznosi X YYY KM" ili "iznosila je X YYY KM"
                # gdje je broj split kao "1 684" ili "1644" ili samo "838"
                # Normaliziraj tekst - zamijeni newline s razmakom za lakse parsiranje
                text_norm = re2.sub(r'\s+', ' ', text)

                patterns_neto = [
                    r"iznosila je (\d{1,2} \d{3}) KM",
                    r"iznosi (\d{1,2} \d{3}) KM",
                    r"amounted to (\d{1,2} \d{3}) KM",
                    r"iznosila je (\d{3,4}) KM",
                    r"iznosi (\d{3,4}) KM",
                    r"amounted to (\d{3,4}) KM",
                    r"iznosila je (\d) (\d{3}) KM",
                    r"amounted to (\d) (\d{3}) KM",
                ]

                for pat in patterns_neto:
                    m = re2.search(pat, text_norm, re2.IGNORECASE)
                    if m:
                        try:
                            if m.lastindex == 2:
                                val = int(m.group(1) + m.group(2))
                            else:
                                val = int(m.group(1).replace(" ", ""))
                            if 700 <= val <= 3000:
                                entry["neto"] = val
                                break
                        except: pass

                # Bruto - slican pattern ali veci iznos
                patterns_bruto = [
                    r"bruto plac[ae][^\d]+(\d{1,2} \d{3}) KM",
                    r"gross wage[^\d]+(\d{1,2} \d{3}) KM",
                ]
                for pat in patterns_bruto:
                    m = re2.search(pat, text, re2.IGNORECASE)
                    if m:
                        try:
                            val = int(m.group(1).replace(" ", ""))
                            if 1000 <= val <= 5000:
                                entry["bruto"] = val
                                break
                        except: pass

                if entry.get("neto"):
                    place_data[period] = entry
                    print(f"  OK {period} ({lang}): neto={entry.get('neto')} KM, bruto={entry.get('bruto','-')} KM")
                    break

            except Exception: pass

    # Sektori iz LAB_01 Excel (ili iz LAB_04 PDF tablice po sektorima)
    sektori = {}
    try:
        xls = fetch_excel("https://bhas.gov.ba/data/Publikacije/VremenskeSerije/LAB_01.xlsx")
        wb = openpyxl.load_workbook(xls, data_only=True)
        for sname in wb.sheetnames[:1]:
            parsed = parse_sheet(wb[sname])
            if parsed:
                sektori[sname] = compact_series(parsed, n_periods=72)
                print(f"  OK Sektori '{sname}': {len(sektori[sname])} serija")
    except Exception as e:
        print(f"  X LAB_01: {e}")

    if not place_data and not sektori:
        path = os.path.join(DATA_DIR, "place.json")
        if not os.path.exists(path):
            save_json("place.json", {"source":"BHAS","error":"Nedostupno","updated":datetime.now().isoformat()[:10],"sheets":{}})
        return False

    sorted_p = sort_periods(list(place_data.keys())) if place_data else []
    save_json("place.json", {
        "source": "BHAS LAB_04/LAB_01",
        "name": "Prosjecne place i place po sektorima BiH",
        "updated": datetime.now().isoformat()[:10],
        "note": "LAB_04=prosjecna placa svih zaposlenih BiH, LAB_01=placa po sektorima",
        "periods": sorted_p,
        "place_bih": {p: place_data[p] for p in sorted_p},
        "sheets": sektori
    })
    return True


def fetch_uino_porezi():
    """Preuzima UINO Excel s prihodima od indirektnih poreza 2004-2026."""
    print("-> Indirektni porezi BiH (UINO)...")
    url = "https://www.uino.gov.ba/portal/wp-content/uploads/10-STATISTIKA/1-Prihodi/Prihodi-UKUPNO-2004-2026-objedinjeni.xlsx"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.uino.gov.ba/portal/bs/statistika/",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }
    try:
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        if len(r.content) < 2000: raise ValueError(f"Premali fajl: {len(r.content)} bytes")
        wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
        print(f"  Sheetovi: {wb.sheetnames}")
        all_data = {}
        for sname in wb.sheetnames[:4]:
            ws = wb[sname]
            parsed = parse_sheet(ws)
            if parsed:
                compact = compact_series(parsed, n_periods=240, max_series=30)
                all_data[sname] = compact
                first = next(iter(compact.values()))
                print(f"  OK '{sname}': {len(compact)} serija, {len(first)} perioda")
        if not all_data: raise ValueError("Nema podataka")
        save_json("uino_porezi.json", {"source":"UINO BiH","name":"Prihodi od indirektnih poreza BiH 2004-2026","url":url,"updated":datetime.now().isoformat()[:10],"note":"PDV, akcize, carine, putarine","sheets":all_data})
        return True
    except Exception as e:
        print(f"  X UINO: {e}")
        path = os.path.join(DATA_DIR, "uino_porezi.json")
        if not os.path.exists(path):
            save_json("uino_porezi.json", {"source":"UINO BiH","error":str(e),"updated":datetime.now().isoformat()[:10],"sheets":{}})
        return False


def fetch_vozila():
    """
    Parsira BHAS TRA_05 PDF saopstenja za registraciju vozila BiH.
    Trazimo broj iz recenice "prvi put registrovano je X motornih vozila"
    """
    print("-> Registracija vozila BiH (BHAS TRA_05)...")
    from datetime import date
    import re as re2
    today = date.today()
    results = {}

    for delta in range(0, 48):
        month = today.month - delta
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        period = str(year) + "-" + str(month)

        for lang in ["BS", "HR"]:
            url = f"https://bhas.gov.ba/data/Publikacije/Saopstenja/{year}/TRA_05_{year}_{month:02d}_1_{lang}.pdf"
            try:
                r = requests.get(url, headers=HEADERS, timeout=12)
                if r.status_code != 200 or len(r.content) < 3000: continue

                with pdfplumber.open(BytesIO(r.content)) as pdf:
                    text = "".join(p.extract_text() or "" for p in pdf.pages[:2])
                if not text: continue

                # Trazi specificnu recenicnu: "prvi put registrovano je X motorn"
                # ili "registrovano je X" ili "registered X motor"
                ukupno = None
                patterns = [
                    r"registrovano je (\d[\d ]+\d) motor",
                    r"registrovana su (\d[\d ]+\d) motor",
                    r"registered (\d[\d ]+\d) motor",
                    r"registrirano je (\d[\d ]+\d) motor",
                ]
                for pat in patterns:
                    m = re2.search(pat, text, re2.IGNORECASE)
                    if m:
                        try:
                            ukupno = int(m.group(1).replace(" ", ""))
                            break
                        except: pass

                if ukupno and 1000 <= ukupno <= 50000:
                    # Trazi udio putnickih automobila
                    putnicki_pct = None
                    m2 = re2.search(r"putnick[^(]+\((\d+\.\d+)%\)", text)
                    if m2:
                        try: putnicki_pct = float(m2.group(1))
                        except: pass

                    results[period] = {
                        "ukupno": ukupno,
                        "putnicki": int(ukupno * putnicki_pct / 100) if putnicki_pct else None,
                        "putnicki_pct": putnicki_pct
                    }
                    print(f"  OK {period} ({lang}): ukupno={ukupno}, putnicki={putnicki_pct}%")
                    break

            except Exception: pass

    if not results:
        print("  Nema podataka iz BHAS TRA_05")
        path = os.path.join(DATA_DIR, "vozila.json")
        if not os.path.exists(path):
            save_json("vozila.json", {"source":"BHAS TRA_05","error":"PDF nije dostupan","updated":datetime.now().isoformat()[:10],"data":{}})
        return False

    sorted_p = sort_periods(list(results.keys()))
    save_json("vozila.json", {"source":"BHAS TRA_05","name":"Registracija motornih vozila BiH","updated":datetime.now().isoformat()[:10],"periods":sorted_p,"data":{p:results[p] for p in sorted_p}})
    return True


def fetch_pufbih_direktni():
    """Parsira PUFBiH saopstenja za direktne poreze FBiH."""
    print("-> Direktni porezi FBiH (PUFBiH)...")
    from datetime import date
    import re as re2
    today = date.today()
    results = {}

    for delta in range(0, 12):
        month = today.month - delta
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        period = str(year) + "-" + str(month)

        urls = [
            f"https://www.pufbih.ba/v1/public/upload/files/Uplate-javnih-prihoda-{month:02d}-{year}.pdf",
            f"https://www.pufbih.ba/v1/public/upload/files/uplate-javnih-prihoda-{month:02d}-{year}.pdf",
        ]
        for url in urls:
            try:
                r = requests.get(url, headers=HEADERS, timeout=8)
                if r.status_code != 200 or len(r.content) < 3000: continue
                with pdfplumber.open(BytesIO(r.content)) as pdf:
                    text = "".join(p.extract_text() or "" for p in pdf.pages[:2])
                if not text: continue

                entry = {}
                for line in text.split("\n"):
                    nums = [n for n in re2.findall(r"[\d,.]+", line) if len(n) > 4]
                    if not nums: continue
                    if any(k in line.lower() for k in ["dohodak", "fizickih"]):
                        for n in nums:
                            try:
                                val = float(n.replace(".", "").replace(",", "."))
                                if 1e7 < val < 2e9: entry["porez_dohodak"] = val; break
                            except: pass
                    if any(k in line.lower() for k in ["dobit", "pravnih"]):
                        for n in nums:
                            try:
                                val = float(n.replace(".", "").replace(",", "."))
                                if 1e7 < val < 2e9: entry["porez_dobit"] = val; break
                            except: pass

                if entry:
                    results[period] = entry
                    print(f"  OK {period}: {entry}")
                    break
            except Exception: pass

    if not results:
        print("  Nema podataka iz PUFBiH")
        path = os.path.join(DATA_DIR, "direktni_porezi.json")
        if not os.path.exists(path):
            save_json("direktni_porezi.json", {"source":"PUFBiH","error":"PDF nije dostupan","updated":datetime.now().isoformat()[:10],"data":{}})
        return False

    sorted_p = sort_periods(list(results.keys()))
    save_json("direktni_porezi.json", {"source":"PUFBiH","name":"Direktni porezi FBiH","updated":datetime.now().isoformat()[:10],"periods":sorted_p,"data":{p:results[p] for p in sorted_p}})
    return True


def update_meta(results):
    meta = {'last_run': datetime.now().isoformat(), 'updated': datetime.now().isoformat()[:10], 'datasets': {}}
    dataset_files = [
        ('maloprodaja', 'maloprodaja.json'), ('turizam', 'turizam.json'),
        ('industrija', 'industrija.json'), ('vanjska_trgovina', 'vanjska_trgovina.json'),
        ('cpi', 'cpi.json'), ('place', 'place.json'),
        ('zaposlenost', 'zaposlenost.json'), ('nezaposlenost', 'nezaposlenost.json'),
        ('uino_porezi', 'uino_porezi.json'), ('vozila', 'vozila.json'),
    ]
    for key, fname in dataset_files:
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            with open(path) as f:
                d = json.load(f)
            meta['datasets'][key] = {
                'name': d.get('name', key), 'updated': d.get('updated'),
                'has_data': bool(d.get('sheets') or d.get('data')),
                'has_error': 'error' in d,
                'size_kb': os.path.getsize(path) // 1024
            }
    save_json('meta.json', meta)



def fetch_cbbh_banke():
    """
    Preuzima CBBH Statistical Appendix Excel fajlove:
    - Attachment 12a: Krediti komercijalnih banaka
    - Attachment 10a: Depoziti kod komercijalnih banaka  
    - Attachment 7a: Krediti domacinstvima po namjeni
    - Attachment 4a: Bilanca komercijalnih banaka
    """
    print("-> Bankarski sektor BiH (CBBH Statistical Appendix)...")

    cbbh_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.cbbh.ba/content/read/1122",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }

    datasets = [
        ("krediti", "Krediti komercijalnih banaka",
         "https://www.cbbh.ba/content/DownloadAttachment/?id=7b5f458b-0d66-49c0-8791-99267bc04ccf&langTag=en"),
        ("depoziti", "Depoziti kod komercijalnih banaka",
         "https://www.cbbh.ba/content/DownloadAttachment/?id=53a74b79-f1c9-49d3-8de7-1df1f86e1f53&langTag=en"),
        ("krediti_domacinstava", "Krediti domacinstvima po namjeni",
         "https://www.cbbh.ba/content/DownloadAttachment/?id=72b09814-d9f1-400f-af5d-04b62f36659d&langTag=en"),
        ("bilanca", "Bilanca komercijalnih banaka",
         "https://www.cbbh.ba/content/DownloadAttachment/?id=8f58ace6-b5f4-419e-bf68-096e05bfcb64&langTag=en"),
        ("dsi", "Direktne strane investicije po zemlji",
         "https://www.cbbh.ba/content/DownloadAttachment/?id=3b165775-0784-4b9a-932f-60be41fcb8ec&langTag=en"),
    ]

    all_sheets = {}

    for key, name, url in datasets:
        try:
            r = requests.get(url, headers=cbbh_headers, timeout=30)
            if r.status_code != 200 or len(r.content) < 2000:
                print(f"  X {key}: HTTP {r.status_code}, {len(r.content)} bytes")
                continue

            wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
            print(f"  OK {key} sheetovi: {wb.sheetnames[:5]}")

            for sname in wb.sheetnames[:3]:
                ws = wb[sname]
                parsed = parse_sheet(ws)
                if parsed and len(parsed) > 0:
                    compact = compact_series(parsed, n_periods=120)
                    if compact:
                        all_sheets[f"{key}_{sname}"] = compact
                        first = next(iter(compact.values()))
                        print(f"    OK '{sname}': {len(compact)} serija, {len(first)} perioda")

        except Exception as e:
            print(f"  X {key}: {e}")

    if not all_sheets:
        print("  Nema podataka s CBBH - URL-ovi blokirani na GitHub serverima")
        path = os.path.join(DATA_DIR, "banke.json")
        if not os.path.exists(path):
            save_json("banke.json", {
                "source": "CBBH", "error": "403 Forbidden",
                "updated": datetime.now().isoformat()[:10], "sheets": {}
            })
        return False

    save_json("banke.json", {
        "source": "CBBH Statistical Appendix",
        "name": "Bankarski sektor BiH",
        "url": "https://www.cbbh.ba/content/read/1122",
        "updated": datetime.now().isoformat()[:10],
        "note": "Krediti, depoziti, bilanca, DSI",
        "sheets": all_sheets
    })
    print(f"  OK banke.json ({len(all_sheets)} dataseta)")
    return True



def update_meta(results):
    meta = {'last_run': datetime.now().isoformat(), 'updated': datetime.now().isoformat()[:10], 'datasets': {}}
    for key in results:
        path = os.path.join(DATA_DIR, key+'.json')
        if os.path.exists(path):
            with open(path) as f:
                d = json.load(f)
            meta['datasets'][key] = {
                'updated': d.get('updated'), 'has_data': bool(d.get('sheets') or d.get('data')),
                'size_kb': os.path.getsize(path) // 1024
            }
    save_json('meta.json', meta)



def fetch_cbbh_banke():
    """
    Preuzima CBBH Statistical Appendix Excel fajlove:
    - Attachment 12a: Krediti komercijalnih banaka
    - Attachment 10a: Depoziti kod komercijalnih banaka
    - Attachment 4a: Bilanca komercijalnih banaka
    """
    print("-> Bankarski sektor BiH (CBBH Statistical Appendix)...")

    cbbh_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.cbbh.ba/",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }

    datasets = {
        "krediti": "https://www.cbbh.ba/content/DownloadAttachment/?id=7b5f458b-0d66-49c0-8791-99267bc04ccf&langTag=en",
        "depoziti": "https://www.cbbh.ba/content/DownloadAttachment/?id=53a74b79-f1c9-49d3-8de7-1df1f86e1f53&langTag=en",
        "krediti_domacinstava": "https://www.cbbh.ba/content/DownloadAttachment/?id=72b09814-d9f1-400f-af5d-04b62f36659d&langTag=en",
        "bilanca": "https://www.cbbh.ba/content/DownloadAttachment/?id=8f58ace6-b5f4-419e-bf68-096e05bfcb64&langTag=en",
    }

    all_sheets = {}

    for name, url in datasets.items():
        try:
            r = requests.get(url, headers=cbbh_headers, timeout=30)
            if r.status_code != 200 or len(r.content) < 2000:
                print(f"  X {name}: HTTP {r.status_code}")
                continue

            wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
            print(f"  {name} sheetovi: {wb.sheetnames[:4]}")

            for sname in wb.sheetnames[:3]:
                parsed = parse_sheet(wb[sname])
                if parsed and len(parsed) >= 2:
                    compact = compact_series(parsed, n_periods=120)
                    if compact:
                        key = f"{name}_{sname}"
                        all_sheets[key] = compact
                        first = next(iter(compact.values()))
                        print(f"    OK '{sname}': {len(compact)} serija, {len(first)} perioda")
                        break  # Uzmi samo prvi korisni sheet

        except Exception as e:
            print(f"  X {name}: {e}")

    if not all_sheets:
        print("  Nema podataka s CBBH")
        path = os.path.join(DATA_DIR, "banke.json")
        if not os.path.exists(path):
            save_json("banke.json", {
                "source": "CBBH", "error": "Nedostupno",
                "updated": datetime.now().isoformat()[:10], "sheets": {}
            })
        return False

    save_json("banke.json", {
        "source": "CBBH Statistical Appendix",
        "name": "Bankarski sektor BiH - krediti, depoziti, bilanca",
        "url": "https://www.cbbh.ba/content/read/1122",
        "updated": datetime.now().isoformat()[:10],
        "note": "Krediti i depoziti komercijalnih banaka BiH",
        "sheets": all_sheets
    })
    return True

if __name__ == '__main__':
    print(f"\n{'='*55}")
    print(f"Makro BiH - Osvjezavanje podataka")
    print(f"Datum: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")
    print(f"{'='*55}\n")

    results = {}
    results['maloprodaja'] = fetch_standard('maloprodaja','Indeksi prometa trgovine na malo',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/STS_01.xlsx'],[0,1],'maloprodaja.json')
    print()
    results['turizam'] = fetch_standard('turizam','Turizam',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/TUR_01.xlsx'],[0,1,2],'turizam.json')
    print()
    results['industrija'] = fetch_standard('industrija','Ind. proizvodnja',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/IND_01.xlsx'],[0,1],'industrija.json')
    print()
    results['vanjska_trgovina'] = fetch_vanjska_trgovina()
    print()
    results['cpi'] = fetch_cpi_pdf()
    print()
    results['bdp'] = fetch_standard('bdp','BDP',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/NAT_01.xlsx',
         'https://bhas.gov.ba/data/Publikacije/VremenskeSerije/NAT_02.xlsx'],[0,1],'bdp.json')
    print()
    results['place'] = fetch_place_neto_bruto()
    print()
    results['zaposlenost'] = fetch_standard('zaposlenost','Zaposleni i trziste rada (ARS)',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/LAB_04.xlsx',
         'https://bhas.gov.ba/data/Publikacije/VremenskeSerije/LAB_02.xlsx',
         'https://bhas.gov.ba/data/Publikacije/VremenskeSerije/LAB_06.xlsx'],
        [0,1,2,3],'zaposlenost.json')
    print()
    results['nezaposlenost'] = fetch_standard('nezaposlenost','Stopa nezaposlenosti (ARS)',
        ['https://bhas.gov.ba/data/Publikacije/VremenskeSerije/LAB_05.xlsx',
         'https://bhas.gov.ba/data/Publikacije/VremenskeSerije/UNE_01.xlsx'],[0,1,2],'nezaposlenost.json')
    print()
    results['uino_porezi'] = fetch_uino_porezi()
    print()
    results['vozila'] = fetch_vozila()
    print()
    results['direktni_porezi'] = fetch_pufbih_direktni()
    print()
    results['banke'] = fetch_cbbh_banke()
    print()

    print("-> Meta...")
    results["banke"] = fetch_cbbh_banke()
    print()
    update_meta(results)

    success = sum(1 for v in results.values() if v)
    print(f"\n{'='*55}")
    print(f"Zavrseno: {success}/{len(results)} uspjesno")
    for key, ok in results.items():
        print(f"  {'OK' if ok else 'X '} {key}")
    print(f"{'='*55}\n")
